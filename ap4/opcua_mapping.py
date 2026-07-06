from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from .diagnostics import Diagnostic, Severity, make_diagnostic
from .fault_catalog import FaultCode


@dataclass(frozen=True)
class OpcUaSignal:
    api_name: str
    canonical_name: str
    required_for_ap4: bool
    description: str
    unit: str | None = None
    node_id: str | None = None


SIGNAL_CONTRACT: tuple[OpcUaSignal, ...] = (
    OpcUaSignal("aktueller_schritt", "aktueller_schritt", True, "SPS-Schritt zur Plausibilisierung"),
    OpcUaSignal("durchfluss_nachguss_maische", "flow_k1_to_k2_l_min", True, "Durchfluss K1 nach K2", "L/min"),

    OpcUaSignal("k1_temperatur", "k1_temperature_c", True, "Temperatur Nachgussbehälter K1", "°C"),
    OpcUaSignal("k1_fuellstand", "k1_level_l", True, "Füllstand Nachgussbehälter K1", "L"),
    OpcUaSignal("k1_maximaler_fuellstand", "k1_level_max", True, "Maximaler Füllstand K1"),
    OpcUaSignal("k1_minimaler_fuellstand", "k1_level_min", True, "Minimaler Füllstand K1"),

    OpcUaSignal("k2_temperatur", "k2_temperature_c", True, "Temperatur Maischebehälter K2", "°C"),
    OpcUaSignal("k2_fuellstand", "k2_level_l", True, "Füllstand Maischebehälter K2", "L"),
    OpcUaSignal("k2_fuellstand_voll", "k2_level_full", True, "Vollmeldung K2"),

    OpcUaSignal("k3_temperatur", "k3_temperature_c", True, "Temperatur Läuter-/Kochbehälter K3", "°C"),
    OpcUaSignal("k3_fuellstand", "k3_level_l", True, "Füllstand K3", "L"),
    OpcUaSignal("k3_fuellstand_voll", "k3_level_full", True, "Vollmeldung K3"),

    OpcUaSignal("mobiler_sensor_temperatur", "k4_temperature_c", True, "Temperatur K4 / Gärtemperatur", "°C"),
)


def canonicalize_values(raw_values: dict[str, Any]) -> dict[str, Any]:
    mapped = dict(raw_values)
    for signal in SIGNAL_CONTRACT:
        if signal.api_name in raw_values:
            mapped[signal.canonical_name] = raw_values[signal.api_name]
    return mapped


def required_api_names() -> list[str]:
    return [s.api_name for s in SIGNAL_CONTRACT if s.required_for_ap4]


def validate_required_signals(raw_values: dict[str, Any]) -> list[Diagnostic]:
    missing = [name for name in required_api_names() if name not in raw_values]
    if not missing:
        return []
    return [
        make_diagnostic(
            Severity.ERROR,
            FaultCode.ERROR_020_AP3_REQUIRED_SIGNAL_MISSING,
            "missing_api_names",
            ",".join(missing),
            "all required api names",
        )
    ]


def load_signal_contract_from_excel(path: str | Path) -> list[OpcUaSignal]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return list(SIGNAL_CONTRACT)

    path = Path(path)
    if not path.exists():
        return list(SIGNAL_CONTRACT)

    workbook = load_workbook(path, data_only=True)
    signals: list[OpcUaSignal] = []
    lower_required = {s.api_name.lower(): s for s in SIGNAL_CONTRACT}

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

        for row in rows[1:]:
            cells = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            api = _first_non_empty(cells, ["api_name", "apiname", "api-name", "signal", "name"])
            node = _first_non_empty(cells, ["nodeid", "node_id", "knotenid", "knoten-id"])
            unit = _first_non_empty(cells, ["einheit", "unit"])

            if api and str(api).lower() in lower_required:
                base = lower_required[str(api).lower()]
                signals.append(
                    OpcUaSignal(
                        base.api_name,
                        base.canonical_name,
                        True,
                        base.description,
                        str(unit) if unit is not None else base.unit,
                        str(node) if node is not None else None,
                    )
                )

    return signals or list(SIGNAL_CONTRACT)


def _first_non_empty(cells: dict[str, Any], names: Iterable[str]) -> Any:
    for name in names:
        value = cells.get(name)
        if value not in (None, ""):
            return value
    return None