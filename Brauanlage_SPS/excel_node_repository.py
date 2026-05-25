from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from signal_model import NodeDefinition, NodePriority


class NodeRepositoryError(RuntimeError):
    pass


class ExcelNodeRepository:
    """Lädt die von AP1 gelieferte OPC-UA-Knotenliste.

    Verwendet gezielt die Blätter:
    - Projektrelevante_Knoten: vollständige Übergabetabelle
    - Top_Pflichtknoten: priorisierte, farblich markierte Pflichtknoten

    Die farbliche Markierung wird nicht als alleinige Wahrheit behandelt, sondern
    als Engineering-Hinweis: hellgelb = Top/Pflicht, blassgelb = Soll-/Grenzwerte,
    orange = Wert/Skalierung prüfen.
    """

    PROJECT_SHEET = "Projektrelevante_Knoten"
    TOP_SHEET = "Top_Pflichtknoten"

    BRIGHT_YELLOW = "FFFFFF00"
    PALE_YELLOW = "FFFFF2CC"
    ORANGE = "FFFCE4D6"

    def __init__(self, workbook_path: str | Path) -> None:
        self.path = Path(workbook_path)
        if not self.path.exists():
            raise NodeRepositoryError(f"Knotenliste nicht gefunden: {self.path}")
        self._nodes: dict[str, NodeDefinition] | None = None

    def load(self) -> dict[str, NodeDefinition]:
        if self._nodes is not None:
            return self._nodes
        wb = load_workbook(self.path, data_only=True)
        if self.PROJECT_SHEET not in wb.sheetnames:
            raise NodeRepositoryError(f"Blatt '{self.PROJECT_SHEET}' fehlt in {self.path.name}")

        project_nodes = self._read_sheet(wb[self.PROJECT_SHEET])
        top_nodes: dict[str, NodeDefinition] = {}
        if self.TOP_SHEET in wb.sheetnames:
            top_nodes = self._read_sheet(wb[self.TOP_SHEET])

        # Top-Blatt überschreibt Priorität/Metadaten, aber nicht die fachliche Basis.
        merged = dict(project_nodes)
        for api_name, top_def in top_nodes.items():
            merged[api_name] = top_def

        self._nodes = merged
        return merged

    def required_for_fsm(self) -> dict[str, NodeDefinition]:
        nodes = self.load()
        return {
            name: nd for name, nd in nodes.items()
            if nd.use_fsm or nd.use_anomaly or nd.priority in {NodePriority.TOP_PFLICHT, NodePriority.PARAMETER_LIMIT, NodePriority.NEEDS_REVIEW}
        }

    def mandatory(self) -> dict[str, NodeDefinition]:
        nodes = self.load()
        return {name: nd for name, nd in nodes.items() if (nd.mandatory_status or "").lower() == "pflicht"}

    def by_api_name(self, api_name: str) -> NodeDefinition:
        nodes = self.load()
        try:
            return nodes[api_name]
        except KeyError as exc:
            raise NodeRepositoryError(f"OPC-UA-Knoten '{api_name}' fehlt in Knotenliste") from exc

    def _read_sheet(self, ws) -> dict[str, NodeDefinition]:
        headers = [cell.value for cell in ws[1]]
        result: dict[str, NodeDefinition] = {}
        for row_idx in range(2, ws.max_row + 1):
            row_values = {headers[i]: ws.cell(row_idx, i + 1).value for i in range(len(headers))}
            api_name = row_values.get("JSON-/API-Name")
            node_id = row_values.get("Node ID")
            if not api_name or not node_id:
                continue
            nd = NodeDefinition(
                original_no=self._as_int(row_values.get("Original #")),
                node_id=str(node_id),
                namespace=self._as_str(row_values.get("Namespace")),
                node_path=self._as_str(row_values.get("Node Path")),
                browse_name=self._as_str(row_values.get("BrowseName")),
                display_name=self._as_str(row_values.get("Display Name")),
                area=self._as_str(row_values.get("Anlagenbereich")),
                component=self._as_str(row_values.get("Komponente")),
                quantity=self._as_str(row_values.get("Mess-/Stellgröße")),
                datatype=self._as_str(row_values.get("Datentyp")),
                category=self._as_str(row_values.get("Kategorie")),
                unit=self._as_str(row_values.get("Einheit")),
                polling_mode=self._as_str(row_values.get("Polling-Modus")),
                polling_interval_s=self._as_float(row_values.get("Polling-Intervall_s")),
                mandatory_status=self._as_str(row_values.get("Pflichtstatus")),
                validation_status=self._as_str(row_values.get("Validierungsstatus")),
                project_access=self._as_str(row_values.get("Zugriffsart_Projekt")),
                use_collector=self._yes(row_values.get("Nutzung_Python_Collector")),
                use_fsm=self._yes(row_values.get("Nutzung_FSM")),
                use_anomaly=self._yes(row_values.get("Nutzung_Anomalie")),
                use_api=self._yes(row_values.get("Nutzung_API")),
                api_name=str(api_name),
                description=self._as_str(row_values.get("Beschreibung / Nutzen")),
                comment=self._as_str(row_values.get("Kommentar / Hinweis")),
                priority=self._row_priority(ws, row_idx),
            )
            result[nd.api_name] = nd
        return result

    def _row_priority(self, ws, row_idx: int) -> NodePriority:
        colors = set()
        for col_idx in range(1, ws.max_column + 1):
            fill = ws.cell(row_idx, col_idx).fill
            if not fill or not fill.fill_type:
                continue
            rgb = fill.fgColor.rgb if fill.fgColor.type == "rgb" else None
            if rgb:
                colors.add(rgb.upper())
        if self.BRIGHT_YELLOW in colors:
            return NodePriority.TOP_PFLICHT
        if self.PALE_YELLOW in colors:
            return NodePriority.PARAMETER_LIMIT
        if self.ORANGE in colors:
            return NodePriority.NEEDS_REVIEW
        return NodePriority.NORMAL

    @staticmethod
    def _yes(value: Any) -> bool:
        return str(value).strip().lower() in {"ja", "yes", "true", "1", "optional"}

    @staticmethod
    def _as_str(value: Any) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text if text else None

    @staticmethod
    def _as_float(value: Any) -> float | None:
        if value in (None, "", "-"):
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _as_int(value: Any) -> int | None:
        if value in (None, "", "-"):
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            return None